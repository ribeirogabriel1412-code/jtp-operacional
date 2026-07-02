"""
Motor Comparativo SAP x PRAXIO — JTP Transportes
Compara saldo de estoque entre SAP (Excel) e PRAXIO (CSV)
e grava apenas os itens com desvio no Supabase.

USO:
  1. Preencha SUPABASE_KEY e GARAGEM_ID abaixo
  2. Coloque os arquivos em:
       SAP:    C:\Users\gabriel.ribeiro\Downloads\3_Valor do Estoque por Dia e Filial.xlsx
       PRAXIO: C:\Users\gabriel.ribeiro\Downloads\PRAXIO.CSV
  3. Execute: python motor_sap_praxio.py

TABELA SUPABASE NECESSÁRIA:
  Execute o SQL em scripts/criar_almox_conferencia_sap_praxio.sql
"""

import pandas as pd
import json, urllib.request, urllib.error
from datetime import datetime, timezone

# ── CONFIGURAÇÕES — preencher antes de rodar ──────────────────────────────────
SUPABASE_URL = 'https://yxwxcxdegkvjvwchemsm.supabase.co'
SUPABASE_KEY = 'COLE_AQUI'       # service_role key (não commitar)
GARAGEM_ID   = 'COLE_GARAGEM_ID' # UUID da garagem Porto Velho

SAP_FILE    = r'C:\Users\gabriel.ribeiro\Downloads\3_Valor do Estoque por Dia e Filial.xlsx'
PRAXIO_FILE = r'C:\Users\gabriel.ribeiro\Downloads\PRAXIO.CSV'
OUT_EXCEL   = r'C:\Users\gabriel.ribeiro\Documents\APP\Relatorios\JTP - Motor SAP x PRAXIO.xlsx'
# ─────────────────────────────────────────────────────────────────────────────

def carregar_sap(path):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df['dep'] = pd.to_numeric(df['Código do depósito'], errors='coerce')
    sap = df[df['dep'] == 200].copy()
    sap['cod']       = sap['Código do item'].str.strip()
    sap['saldo_sap'] = pd.to_numeric(sap['Saldo em estoque'], errors='coerce').fillna(0)
    sap['preco_sap'] = pd.to_numeric(sap['Preço médio'], errors='coerce').fillna(0)
    sap['valor_sap'] = pd.to_numeric(sap['Valor em estoque'], errors='coerce').fillna(0)
    data_ref = sap['Data'].dropna().iloc[0] if not sap['Data'].dropna().empty else ''
    return sap, str(data_ref)


def carregar_praxio(path):
    """
    O CSV do PRAXIO tem 17 campos por linha (número seq. no início + 15 dados + trailing empty).
    O saldo real de estoque está na posição 13 (campo "Saldo_qtd" neste mapeamento).
    """
    COLS = [
        'Seq', 'Empresa', 'Filial', 'Local', 'Desc_local',
        'Cod_grupo', 'Desc_grupo',
        'Cod_interno', 'Descr_item', 'Classe', 'Subclasse', 'Localizacao',
        'Minimo', 'Saldo_qtd',      # posição 13 = saldo real
        'Media_preco',               # posição 14 = preço médio
        'Total',                     # posição 15 = valor total
        'Extra'
    ]
    df = pd.read_csv(
        path, sep=',', quotechar='"', encoding='latin-1',
        dtype=str, names=COLS, header=0
    )
    df['cod']          = df['Cod_interno'].str.strip()
    df['saldo_praxio'] = pd.to_numeric(df['Saldo_qtd'], errors='coerce').fillna(0)
    df['preco_praxio'] = pd.to_numeric(
        df['Media_preco'].str.replace(',', '.', regex=False), errors='coerce'
    ).fillna(0)
    return df


def comparar(sap, praxio):
    merged = pd.merge(
        sap[['cod', 'Descrição do item', 'Nome do grupo', 'saldo_sap', 'preco_sap', 'valor_sap']],
        praxio[['cod', 'Descr_item', 'saldo_praxio', 'preco_praxio']],
        on='cod', how='outer'
    )
    merged['saldo_sap']    = merged['saldo_sap'].fillna(0)
    merged['saldo_praxio'] = merged['saldo_praxio'].fillna(0)
    merged['diferenca']    = merged['saldo_sap'] - merged['saldo_praxio']
    merged['descricao']    = merged['Descrição do item'].fillna(merged['Descr_item'])
    merged['grupo']        = merged['Nome do grupo'].fillna('')
    desvios = merged[merged['diferenca'] != 0].copy()
    desvios = desvios.sort_values('diferenca', key=abs, ascending=False).reset_index(drop=True)
    return merged, desvios


def exportar_excel(merged, desvios, data_ref, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill  = PatternFill('solid', start_color='1E3A5F')
    amber_fill   = PatternFill('solid', start_color='F59E0B')
    red_fill     = PatternFill('solid', start_color='EF4444')
    white_bold   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    black_normal = Font(name='Arial', size=10)
    center       = Alignment(horizontal='center', vertical='center')
    left_m       = Alignment(horizontal='left', vertical='center')
    thin         = Side(style='thin', color='CCCCCC')
    bdr          = Border(top=thin, bottom=thin, left=thin, right=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Desvios SAP x PRAXIO'

    headers = [
        ('#', 5), ('Cód. SAP', 13), ('Descrição', 48), ('Grupo', 22),
        ('Saldo SAP', 12), ('Saldo PRAXIO', 14), ('Diferença', 12),
        ('Preço SAP', 12), ('Valor SAP', 14), ('Status', 16)
    ]
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = white_bold
        c.fill = header_fill
        c.alignment = center
        c.border = bdr
        ws.column_dimensions[c.column_letter].width = w

    for ri, row in desvios.iterrows():
        r = ri + 2
        dif = row['diferenca']
        status = ('Só no SAP'    if row['saldo_praxio'] == 0 and row['saldo_sap'] > 0
                  else 'Só no PRAXIO' if row['saldo_sap'] == 0 and row['saldo_praxio'] > 0
                  else 'Divergência')
        vals = [
            ri + 1, row['cod'], row['descricao'], row['grupo'],
            row['saldo_sap'], row['saldo_praxio'], dif,
            row['preco_sap'] if pd.notna(row.get('preco_sap')) else 0,
            row['valor_sap'] if pd.notna(row.get('valor_sap')) else 0,
            status
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.font = black_normal
            c.border = bdr
            c.alignment = center if ci in (1, 5, 6, 7, 8, 9) else left_m
            if ci == 7:
                c.font = Font(name='Arial', size=10, bold=True,
                              color='EF4444' if dif < 0 else '00C48C')
            if ci == 10:
                c.fill = amber_fill if status == 'Divergência' else red_fill
                c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:J{len(desvios)+1}'

    ws2 = wb.create_sheet('Resumo')
    ws2['A1'] = 'Motor Comparativo SAP x PRAXIO'
    ws2['A1'].font = Font(name='Arial', size=14, bold=True)
    resumo = [
        ('A3', 'B3', 'Data referência (SAP):',           str(data_ref)),
        ('A4', 'B4', 'Gerado em:',                       datetime.now().strftime('%d/%m/%Y %H:%M')),
        ('A6', 'B6', 'Total de itens comparados:',       len(merged)),
        ('A7', 'B7', 'Itens com desvio:',                len(desvios)),
        ('A8', 'B8', 'Só no SAP (sem reg. PRAXIO):',     int((desvios['saldo_praxio']==0).sum())),
        ('A9', 'B9', 'Só no PRAXIO (zerado no SAP):',    int((desvios['saldo_sap']==0).sum())),
        ('A10','B10','Divergências reais (ambos > 0):',   int(((desvios['saldo_sap']>0)&(desvios['saldo_praxio']>0)).sum())),
    ]
    for la, lb, label, val in resumo:
        ws2[la] = label
        ws2[la].font = Font(name='Arial', bold=True, size=10)
        ws2[lb] = val
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 20

    wb.save(path)
    print(f'Excel salvo em: {path}')


def upload_supabase(desvios, data_ref):
    now_iso  = datetime.now(timezone.utc).isoformat()
    data_iso = str(data_ref).replace('/', '-')[:10] if data_ref else now_iso[:10]

    # Remove registros anteriores da mesma data
    url = f'{SUPABASE_URL}/rest/v1/almox_conferencia_sap_praxio'
    req = urllib.request.Request(
        f'{url}?garagem_id=eq.{GARAGEM_ID}&data_referencia=eq.{data_iso}',
        method='DELETE',
        headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}',
                 'Content-Type': 'application/json'}
    )
    try:
        urllib.request.urlopen(req)
    except Exception:
        pass

    records = [{
        'garagem_id':      GARAGEM_ID,
        'data_referencia': data_iso,
        'cod_sap':         str(r['cod']),
        'descricao':       str(r['descricao'] or '')[:200],
        'grupo':           str(r['grupo'] or '')[:100],
        'saldo_sap':       float(r['saldo_sap']),
        'saldo_praxio':    float(r['saldo_praxio']),
        'diferenca':       float(r['diferenca']),
        'preco_medio_sap': float(r['preco_sap'] if pd.notna(r.get('preco_sap')) else 0),
        'valor_sap':       float(r['valor_sap'] if pd.notna(r.get('valor_sap')) else 0),
        'criado_em':       now_iso,
    } for _, r in desvios.iterrows()]

    total_ok = 0
    for i in range(0, len(records), 200):
        batch = records[i:i+200]
        req = urllib.request.Request(
            url, data=json.dumps(batch).encode('utf-8'), method='POST',
            headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}',
                     'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
        )
        try:
            urllib.request.urlopen(req)
            total_ok += len(batch)
            print(f'  Lote {i//200+1}: {len(batch)} registros inseridos.')
        except urllib.error.HTTPError as e:
            print(f'  ERRO lote {i//200+1}: {e.code} {e.read().decode()}')

    print(f'Total inserido: {total_ok} desvios.')


if __name__ == '__main__':
    print('Carregando SAP...')
    sap, data_ref = carregar_sap(SAP_FILE)
    print(f'  {len(sap)} itens (depósito 200) | data: {data_ref}')

    print('Carregando PRAXIO...')
    praxio = carregar_praxio(PRAXIO_FILE)
    print(f'  {len(praxio)} itens')

    print('Comparando...')
    merged, desvios = comparar(sap, praxio)
    print(f'  Total: {len(merged)} | Desvios: {len(desvios)}')
    print()
    print(desvios[['cod','descricao','saldo_sap','saldo_praxio','diferenca']].head(20).to_string())

    print('\nGerando Excel...')
    exportar_excel(merged, desvios, data_ref, OUT_EXCEL)

    if SUPABASE_KEY == 'COLE_AQUI':
        print('\n[AVISO] Configure SUPABASE_KEY e GARAGEM_ID para ativar o upload.')
    else:
        print('\nUpload Supabase...')
        upload_supabase(desvios, data_ref)
